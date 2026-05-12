from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

from noboom_benchmark.noboom_lib.core.tune.report_generation import (
    excel_score_table_has_metrics,
    update_score_excel_one,
)


def test_update_score_excel_one_migrates_legacy_best_columns_to_theory_best(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "scores.xlsx"
    df = pd.DataFrame(
        [[3.5, 4.5]],
        index=pd.Index(["legacy_model"], name="Model"),
        columns=pd.MultiIndex.from_tuples(
            [
                ("cont_reactive_ome", "alarm_score"),
                ("cont_reactive_ome", "best_alarm_score"),
            ],
            names=["Dataset", "metrics"],
        ),
    )
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, sheet_name="scores")

    update_score_excel_one(
        "new_model",
        "cont_reactive_ome",
        {
            "alarm_score": 4.0,
            "best_alarm_score": 4.2,
            "theory_best_alarm_score": 5.0,
        },
        excel_path,
        "scores",
    )

    updated = pd.read_excel(
        excel_path,
        sheet_name="scores",
        header=[0, 1],
        index_col=0,
    )

    assert ("cont_reactive_ome", "best_alarm_score") in updated.columns
    assert ("cont_reactive_ome", "theory_best_alarm_score") in updated.columns
    assert pd.isna(updated.loc["legacy_model", ("cont_reactive_ome", "best_alarm_score")])
    assert updated.loc["legacy_model", ("cont_reactive_ome", "theory_best_alarm_score")] == 4.5
    assert updated.loc["new_model", ("cont_reactive_ome", "best_alarm_score")] == 4.2
    assert updated.loc["new_model", ("cont_reactive_ome", "theory_best_alarm_score")] == 5.0


def test_update_score_excel_one_migrates_upper_bound_columns_to_theory_best(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "scores.xlsx"
    df = pd.DataFrame(
        [[3.5, 4.5]],
        index=pd.Index(["legacy_model"], name="Model"),
        columns=pd.MultiIndex.from_tuples(
            [
                ("cont_reactive_ome", "alarm_score"),
                ("cont_reactive_ome", "upper_bound_alarm_score"),
            ],
            names=["Dataset", "metrics"],
        ),
    )
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, sheet_name="scores")

    update_score_excel_one(
        "new_model",
        "cont_reactive_ome",
        {
            "alarm_score": 4.0,
            "best_alarm_score": 4.2,
            "theory_best_alarm_score": 5.0,
        },
        excel_path,
        "scores",
    )

    updated = pd.read_excel(
        excel_path,
        sheet_name="scores",
        header=[0, 1],
        index_col=0,
    )

    assert ("cont_reactive_ome", "theory_best_alarm_score") in updated.columns
    assert updated.loc["legacy_model", ("cont_reactive_ome", "theory_best_alarm_score")] == 4.5
    assert updated.loc["new_model", ("cont_reactive_ome", "theory_best_alarm_score")] == 5.0


def test_update_score_excel_one_bolds_best_metric_columns(
    tmp_path: Path,
) -> None:
    excel_path = tmp_path / "scores.xlsx"

    update_score_excel_one(
        "model_a",
        "cont_reactive_ome",
        {
            "alarm_score": 4.0,
            "best_alarm_score": 4.2,
        },
        excel_path,
        "scores",
    )
    update_score_excel_one(
        "model_b",
        "cont_reactive_ome",
        {
            "alarm_score": 3.5,
            "best_alarm_score": 4.8,
        },
        excel_path,
        "scores",
    )

    updated = pd.read_excel(
        excel_path,
        sheet_name="scores",
        header=[0, 1],
        index_col=0,
    )
    workbook = load_workbook(excel_path)
    worksheet = workbook["scores"]

    col_pos = updated.columns.get_loc(("cont_reactive_ome", "best_alarm_score"))
    excel_col = 2 + col_pos
    model_a_row = 4 + updated.index.get_loc("model_a")
    model_b_row = 4 + updated.index.get_loc("model_b")

    assert not bool(worksheet.cell(row=model_a_row, column=excel_col).font.bold)
    assert bool(worksheet.cell(row=model_b_row, column=excel_col).font.bold)


def test_excel_score_table_has_metrics_detects_missing_columns(tmp_path: Path) -> None:
    excel_path = tmp_path / "scores.xlsx"

    update_score_excel_one("model_a", "ome", {}, excel_path, "scores")

    assert excel_score_table_has_metrics(excel_path, "scores") is False


def test_excel_score_table_has_metrics_detects_existing_metric_columns(tmp_path: Path) -> None:
    excel_path = tmp_path / "scores.xlsx"

    update_score_excel_one("model_a", "ome", {"alarm_score": 1.0}, excel_path, "scores")

    assert excel_score_table_has_metrics(excel_path, "scores") is True
