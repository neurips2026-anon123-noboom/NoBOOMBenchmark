from __future__ import annotations

from pathlib import Path
from types import SimpleNamespace
from typing import Any, Optional

from openpyxl import load_workbook

from noboom_cluster.noboom_cli_lib.specs import PairResult, PairRunSpec

from noboom_benchmark.noboom_lib.core.tune import callbacks
from noboom_benchmark.noboom_lib.core.tune.report_generation import update_score_excel_one


def _build_pair_spec(tmp_path: Path) -> PairRunSpec:
    return PairRunSpec(
        experiment_id="1",
        source_experiment_id=None,
        model_name="gdn",
        dataset_name="ome",
        timestamp="20260311_120000",
        storage_path=str(tmp_path / "storage"),
        gpus_per_run=0.25,
        optuna_storage_uri="postgresql+psycopg://noboom:example-password@127.0.0.1:5432/optuna_db",
        config_dir="configs",
        temp_dir=str(tmp_path / "temp"),
        verbose=0,
        tune=True,
        env_file=None,
        tracking_uri="http://127.0.0.1:5000",
        s3_endpoint_url="http://127.0.0.1:8333",
        prepared_dataset_s3_path=None,
    )


def _build_study_run() -> SimpleNamespace:
    return SimpleNamespace(
        info=SimpleNamespace(
            artifact_uri="s3://noboom-ray/artifacts/study-run",
            run_name="study__gdn__ome__20260311_120000",
        )
    )


def _build_child_run() -> SimpleNamespace:
    return SimpleNamespace(
        info=SimpleNamespace(
            artifact_uri="s3://noboom-ray/artifacts/child-run",
            run_name="trial__001",
        )
    )


def test_sync_pair_outputs_skips_checkpoint_sync_by_default(
    monkeypatch,
    tmp_path: Path,
) -> None:
    experiment_root = tmp_path / "experiment"
    experiment_root.mkdir()
    temp_root = tmp_path / "temp"
    temp_root.mkdir()
    excel_path = experiment_root / "noboom_experiments.xlsx"
    pair_spec = _build_pair_spec(tmp_path)

    callback = callbacks.MlflowSeafilePairOutputCallback(
        experiment_root=experiment_root,
        experiment_name="NoBoomBenchmark__20260311_120000",
        runtime_config=SimpleNamespace(seafile_upload_checkpoints=False),
        config_dir="configs",
        temp_dir=str(temp_root),
        tracking_uri="http://127.0.0.1:5000",
    )

    upload_calls: list[dict[str, Any]] = []
    subprocess_calls: list[list[str]] = []

    def fake_upload(
        storage_path: str | Path,
        remote_storage_path: str | Path,
        extra_args: Optional[list] = None,
        compressed: bool = False,
        archive_name: Optional[str] = None,
    ) -> None:
        upload_calls.append(
            {
                "storage_path": str(storage_path),
                "remote_storage_path": str(remote_storage_path),
                "extra_args": extra_args,
                "compressed": compressed,
                "archive_name": archive_name,
            }
        )

    def fake_subprocess_run(command: list[str], check: bool) -> None:
        assert check is True
        subprocess_calls.append(command)

    monkeypatch.setattr(callbacks, "upload_to_seafile", fake_upload)
    monkeypatch.setattr(callbacks, "gather_runs", lambda *_args, **_kwargs: {"child-run-id": "nested/child"})
    monkeypatch.setattr(callbacks.subprocess, "run", fake_subprocess_run)
    monkeypatch.setattr(callbacks.mlflow, "set_tracking_uri", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(
        callbacks.mlflow,
        "get_run",
        lambda run_id: _build_study_run() if run_id == "study-run-id" else _build_child_run(),
    )
    monkeypatch.setattr(
        callbacks.mlflow,
        "get_experiment_by_name",
        lambda *_args, **_kwargs: SimpleNamespace(experiment_id="123"),
    )

    callback._sync_pair_outputs(pair_spec, "study-run-id", excel_path)  # noqa: SLF001

    assert all("BEST_" not in call["remote_storage_path"] for call in upload_calls)
    assert any(
        call["remote_storage_path"] == "NoBoomBenchmark__20260311_120000"
        and call["extra_args"] == ["--include", "*.xlsx", "--include", "*.json"]
        for call in upload_calls
    )
    assert any(
        call["remote_storage_path"] == "NoBoomBenchmark__20260311_120000/all"
        and call["compressed"] is True
        and call["archive_name"] == "study__gdn__ome__20260311_120000"
        for call in upload_calls
    )
    assert len(subprocess_calls) == 1
    assert "--exclude" in subprocess_calls[0]
    assert "'**/*.ckpt'" in subprocess_calls[0]


def test_sync_pair_outputs_uploads_checkpoints_when_enabled(
    monkeypatch,
    tmp_path: Path,
) -> None:
    experiment_root = tmp_path / "experiment"
    experiment_root.mkdir()
    temp_root = tmp_path / "temp"
    temp_root.mkdir()
    excel_path = experiment_root / "noboom_experiments.xlsx"
    pair_spec = _build_pair_spec(tmp_path)

    callback = callbacks.MlflowSeafilePairOutputCallback(
        experiment_root=experiment_root,
        experiment_name="NoBoomBenchmark__20260311_120000",
        runtime_config=SimpleNamespace(seafile_upload_checkpoints=True),
        config_dir="configs",
        temp_dir=str(temp_root),
        tracking_uri="http://127.0.0.1:5000",
    )

    upload_calls: list[dict[str, Any]] = []

    def fake_upload(
        storage_path: str | Path,
        remote_storage_path: str | Path,
        extra_args: Optional[list] = None,
        compressed: bool = False,
        archive_name: Optional[str] = None,
    ) -> None:
        upload_calls.append(
            {
                "storage_path": str(storage_path),
                "remote_storage_path": str(remote_storage_path),
                "extra_args": extra_args,
                "compressed": compressed,
                "archive_name": archive_name,
            }
        )

    monkeypatch.setattr(callbacks, "upload_to_seafile", fake_upload)
    monkeypatch.setattr(callbacks, "gather_runs", lambda *_args, **_kwargs: {})
    monkeypatch.setattr(callbacks.subprocess, "run", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(callbacks.mlflow, "set_tracking_uri", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(callbacks.mlflow, "get_run", lambda *_args, **_kwargs: _build_study_run())
    monkeypatch.setattr(callbacks.mlflow, "get_experiment_by_name", lambda *_args, **_kwargs: None)

    callback._sync_pair_outputs(pair_spec, "study-run-id", excel_path)  # noqa: SLF001

    assert any(
        call["remote_storage_path"]
        == "NoBoomBenchmark__20260311_120000/BEST_study__gdn__ome__20260311_120000"
        and call["storage_path"] == "seaweed_s3:noboom-ray/artifacts/study-run"
        for call in upload_calls
    )


def test_on_job_terminal_marks_stopped_result_and_skips_cleanup(
    monkeypatch,
    tmp_path: Path,
) -> None:
    experiment_root = tmp_path / "experiment"
    experiment_root.mkdir()
    temp_root = tmp_path / "temp"
    temp_root.mkdir()
    pair_spec = _build_pair_spec(tmp_path)
    storage_root = Path(pair_spec.storage_path)
    storage_root.mkdir(parents=True, exist_ok=True)
    (storage_root / "result.json").write_text(
        PairResult(
            model_name=pair_spec.model_name,
            dataset_name=pair_spec.dataset_name,
            storage_path=pair_spec.storage_path,
            study_run_id="study-run-id",
            result={"alarm_score": 0.9, "best_alarm_score": 0.95},
            status="RUNNING",
            partial_result=True,
            result_source="incremental_best_trial",
        ).model_dump_json(indent=2),
        encoding="utf-8",
    )

    callback = callbacks.MlflowSeafilePairOutputCallback(
        experiment_root=experiment_root,
        experiment_name="NoBoomBenchmark__20260311_120000",
        runtime_config=SimpleNamespace(
            seafile_username="user",
            seafile_root_path="results",
            seafile_upload_results=True,
            seafile_upload_checkpoints=True,
        ),
        config_dir="src/noboom_cluster/cluster_files/configs",
        temp_dir=str(temp_root),
        tracking_uri="http://127.0.0.1:5000",
    )

    sync_calls = []
    cleanup_calls = []

    monkeypatch.setattr(callbacks, "load_json_artifact_from_mlflow", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(
        callback,
        "_sync_pair_outputs",
        lambda *args, **kwargs: sync_calls.append((args, kwargs)),
    )
    monkeypatch.setattr(
        callback,
        "_cleanup_pair_artifacts",
        lambda *args, **kwargs: cleanup_calls.append((args, kwargs)),
    )

    pair_result = callback.on_job_terminal(pair_spec, "STOPPED", "Stopped by user")

    assert pair_result.status == "STOPPED"
    assert pair_result.partial_result is True
    assert pair_result.job_status_message == "Stopped by user"
    assert pair_result.result_source == "incremental_best_trial"
    assert pair_result.seafile_synced is True
    assert pair_result.cleanup_performed is False
    assert len(sync_calls) == 1
    assert sync_calls[0][1]["include_checkpoints"] is False
    assert sync_calls[0][1]["include_run_artifacts"] is False
    assert cleanup_calls == []


def test_on_job_terminal_local_artifact_backend_never_syncs_to_seafile(
    monkeypatch,
    tmp_path: Path,
) -> None:
    experiment_root = tmp_path / "experiment"
    experiment_root.mkdir()
    temp_root = tmp_path / "temp"
    temp_root.mkdir()
    pair_spec = _build_pair_spec(tmp_path).model_copy(
        update={
            "execution_backend": "local",
            "artifact_storage_backend": "local",
            "s3_endpoint_url": None,
        }
    )
    storage_root = Path(pair_spec.storage_path)
    storage_root.mkdir(parents=True, exist_ok=True)
    (storage_root / "result.json").write_text(
        PairResult(
            model_name=pair_spec.model_name,
            dataset_name=pair_spec.dataset_name,
            storage_path=pair_spec.storage_path,
            study_run_id="study-run-id",
            result={"alarm_score": 0.9, "best_alarm_score": 0.95},
            status="SUCCEEDED",
            result_source="local_snapshot",
        ).model_dump_json(indent=2),
        encoding="utf-8",
    )

    callback = callbacks.MlflowSeafilePairOutputCallback(
        experiment_root=experiment_root,
        experiment_name="NoBoomBenchmark__20260311_120000",
        runtime_config=SimpleNamespace(
            seafile_username="user-from-env",
            seafile_root_path="results",
            seafile_upload_checkpoints=True,
            seafile_upload_results=True,
            execution_backend="local",
            artifact_storage_backend="local",
        ),
        config_dir="src/noboom_cluster/cluster_files/configs",
        temp_dir=str(temp_root),
        tracking_uri="http://127.0.0.1:5000",
    )

    monkeypatch.setattr(callbacks, "load_json_artifact_from_mlflow", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(
        callback,
        "_sync_pair_outputs",
        lambda *args, **kwargs: (_ for _ in ()).throw(
            AssertionError("local artifact backend should not sync to Seafile")
        ),
    )
    monkeypatch.setattr(
        callback,
        "_cleanup_pair_artifacts",
        lambda *args, **kwargs: (_ for _ in ()).throw(
            AssertionError("local artifact backend should not clean remote artifacts")
        ),
    )

    pair_result = callback.on_job_terminal(pair_spec, "SUCCEEDED")

    assert pair_result.status == "SUCCEEDED"
    assert pair_result.seafile_synced is False
    assert pair_result.cleanup_performed is False
    assert (experiment_root / "noboom_experiments.xlsx").exists()


def test_on_job_terminal_respects_disabled_seafile_results_upload(
    monkeypatch,
    tmp_path: Path,
) -> None:
    experiment_root = tmp_path / "experiment"
    experiment_root.mkdir()
    temp_root = tmp_path / "temp"
    temp_root.mkdir()
    pair_spec = _build_pair_spec(tmp_path)
    storage_root = Path(pair_spec.storage_path)
    storage_root.mkdir(parents=True, exist_ok=True)
    (storage_root / "result.json").write_text(
        PairResult(
            model_name=pair_spec.model_name,
            dataset_name=pair_spec.dataset_name,
            storage_path=pair_spec.storage_path,
            study_run_id="study-run-id",
            result={"alarm_score": 0.9, "best_alarm_score": 0.95},
            status="SUCCEEDED",
            result_source="local_snapshot",
        ).model_dump_json(indent=2),
        encoding="utf-8",
    )

    callback = callbacks.MlflowSeafilePairOutputCallback(
        experiment_root=experiment_root,
        experiment_name="NoBoomBenchmark__20260311_120000",
        runtime_config=SimpleNamespace(
            seafile_username="user",
            seafile_root_path="results",
            seafile_upload_results=False,
            seafile_upload_checkpoints=True,
        ),
        config_dir="src/noboom_cluster/cluster_files/configs",
        temp_dir=str(temp_root),
        tracking_uri="http://127.0.0.1:5000",
    )

    monkeypatch.setattr(callbacks, "load_json_artifact_from_mlflow", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(
        callback,
        "_sync_pair_outputs",
        lambda *args, **kwargs: (_ for _ in ()).throw(
            AssertionError("disabled Seafile result upload should not sync")
        ),
    )

    pair_result = callback.on_job_terminal(pair_spec, "SUCCEEDED")

    assert pair_result.seafile_synced is False
    assert (experiment_root / "noboom_experiments.xlsx").exists()


def test_on_job_terminal_requires_seafile_root_for_result_sync(
    monkeypatch,
    tmp_path: Path,
) -> None:
    experiment_root = tmp_path / "experiment"
    experiment_root.mkdir()
    temp_root = tmp_path / "temp"
    temp_root.mkdir()
    pair_spec = _build_pair_spec(tmp_path)
    storage_root = Path(pair_spec.storage_path)
    storage_root.mkdir(parents=True, exist_ok=True)
    (storage_root / "result.json").write_text(
        PairResult(
            model_name=pair_spec.model_name,
            dataset_name=pair_spec.dataset_name,
            storage_path=pair_spec.storage_path,
            study_run_id="study-run-id",
            result={"alarm_score": 0.9, "best_alarm_score": 0.95},
            status="SUCCEEDED",
            result_source="local_snapshot",
        ).model_dump_json(indent=2),
        encoding="utf-8",
    )

    callback = callbacks.MlflowSeafilePairOutputCallback(
        experiment_root=experiment_root,
        experiment_name="NoBoomBenchmark__20260311_120000",
        runtime_config=SimpleNamespace(
            seafile_username="user-needed-for-dataset",
            seafile_root_path="",
            seafile_upload_results=True,
            seafile_upload_checkpoints=True,
        ),
        config_dir="src/noboom_cluster/cluster_files/configs",
        temp_dir=str(temp_root),
        tracking_uri="http://127.0.0.1:5000",
    )

    monkeypatch.setattr(callbacks, "load_json_artifact_from_mlflow", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(
        callback,
        "_sync_pair_outputs",
        lambda *args, **kwargs: (_ for _ in ()).throw(
            AssertionError("Seafile sync should require SEAFILE_ROOT_PATH")
        ),
    )

    pair_result = callback.on_job_terminal(pair_spec, "SUCCEEDED")

    assert pair_result.seafile_synced is False
    assert (experiment_root / "noboom_experiments.xlsx").exists()


def test_on_job_terminal_writes_head_excel_status_for_failed_pair(
    monkeypatch,
    tmp_path: Path,
) -> None:
    experiment_root = tmp_path / "experiment"
    experiment_root.mkdir()
    temp_root = tmp_path / "temp"
    temp_root.mkdir()
    pair_spec = _build_pair_spec(tmp_path)

    callback = callbacks.MlflowSeafilePairOutputCallback(
        experiment_root=experiment_root,
        experiment_name="NoBoomBenchmark__20260311_120000",
        runtime_config=SimpleNamespace(seafile_username="", seafile_upload_results=True),
        config_dir="src/noboom_cluster/cluster_files/configs",
        temp_dir=str(temp_root),
        tracking_uri="http://127.0.0.1:5000",
    )

    monkeypatch.setattr(callbacks, "load_json_artifact_from_mlflow", lambda *_args, **_kwargs: None)

    pair_result = callback.on_job_terminal(pair_spec, "FAILED", "worker failed")

    excel_path = experiment_root / "noboom_experiments.xlsx"
    workbook = load_workbook(excel_path)
    values = [
        cell
        for row in workbook["scores"].iter_rows(values_only=True)
        for cell in row
        if cell is not None
    ]

    assert pair_result.status == "FAILED"
    assert pair_result.partial_result is True
    assert "failed" in values


def test_on_job_terminal_adds_blank_row_for_stopped_pair_when_excel_has_metrics(
    monkeypatch,
    tmp_path: Path,
) -> None:
    experiment_root = tmp_path / "experiment"
    experiment_root.mkdir()
    temp_root = tmp_path / "temp"
    temp_root.mkdir()
    excel_path = experiment_root / "noboom_experiments.xlsx"
    update_score_excel_one(
        "model_a",
        "ome",
        {"alarm_score": 0.5},
        excel_path,
        "scores",
    )

    pair_spec = _build_pair_spec(tmp_path)
    callback = callbacks.MlflowSeafilePairOutputCallback(
        experiment_root=experiment_root,
        experiment_name="NoBoomBenchmark__20260311_120000",
        runtime_config=SimpleNamespace(seafile_username="", seafile_upload_checkpoints=False),
        config_dir="src/noboom_cluster/cluster_files/configs",
        temp_dir=str(temp_root),
        tracking_uri="http://127.0.0.1:5000",
    )

    monkeypatch.setattr(callbacks, "load_json_artifact_from_mlflow", lambda *_args, **_kwargs: None)

    pair_result = callback.on_job_terminal(pair_spec, "STOPPED", "Stopped by user")

    assert pair_result.result == {}

    import pandas as pd

    updated = pd.read_excel(
        excel_path,
        sheet_name="scores",
        header=[0, 1],
        index_col=0,
    )
    assert "gdn" in updated.index
    assert updated.loc["gdn"].isna().all()
