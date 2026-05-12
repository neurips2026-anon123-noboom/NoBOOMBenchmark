import logging
from numbers import Number
from pathlib import Path
from typing import Any, Iterable, Mapping, Optional

from filelock import FileLock
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter
import pandas as pd

logger = logging.getLogger(__name__)


def excel_score_table_has_metrics(excel_path: Path, sheet_name: str) -> bool:
    if not excel_path.exists():
        return False

    try:
        frame = pd.read_excel(
            excel_path,
            sheet_name=sheet_name,
            header=[0, 1],
            index_col=0,
        )
    except Exception as exc:  # noqa: BLE001
        logger.warning(
            "Failed to inspect Excel score table '%s' for metric columns: %s",
            excel_path,
            exc,
        )
        return False

    return isinstance(frame.columns, pd.MultiIndex) and len(frame.columns) > 0


def update_score_excel_one(
    model_name: str,
    dataset_name: str,
    metrics: Mapping[str, Any],
    excel_path: Path,
    sheet_name: str,
    minimize_metrics: Optional[Iterable[str]] = None,
) -> None:
    """Update an Excel score table for a model/dataset pair.

    Args:
        model_name (str): Model name whose row should be updated.
        dataset_name (str): Dataset name whose column group should be updated.
        metrics (Mapping[str, float]): Mapping of metric names to values.
        excel_path (Path): Path to the Excel file to update.
        sheet_name (str): Sheet name to read/write.
        minimize_metrics (Optional[Iterable[str]]): Metrics to minimize (bold min).
            Defaults to None.

    Returns:
        None: The Excel file is updated in place.

    Side Effects:
        Reads and writes the Excel file and acquires a file lock.
    """

    minimize_set = set(minimize_metrics or [])
    logger.info(
        "Updating Excel scores: model='%s', dataset='%s', excel_path='%s', sheet='%s', minimize_metrics=%s",
        model_name,
        dataset_name,
        excel_path,
        sheet_name,
        minimize_set,
    )
    logger.debug("Raw metrics provided: %s", metrics)

    lock_path = excel_path.with_suffix(excel_path.suffix + ".lock")
    lock = FileLock(str(lock_path))
    logger.debug("Using file lock at '%s'.", lock_path)

    with lock:
        logger.info("Acquired file lock for '%s'.", excel_path)

        # --- 1) Load current table (if exists) ---
        if excel_path.exists():
            logger.info("Excel file exists; loading workbook into DataFrame.")
            df = pd.read_excel(
                excel_path,
                sheet_name=sheet_name,
                header=[0, 1],   # MultiIndex columns: (Dataset, metrics)
                index_col=0,     # Model
            )
            logger.debug(
                "Loaded DataFrame with shape %s and columns: %s",
                df.shape,
                df.columns,
            )
        else:
            logger.info("Excel file does not exist; starting with empty DataFrame.")
            df = pd.DataFrame()

        # Ensure MultiIndex columns if there is existing data
        if len(df.columns) > 0 and not isinstance(df.columns, pd.MultiIndex):
            logger.warning(
                "Existing DataFrame columns are not MultiIndex; converting to MultiIndex."
            )
            df.columns = pd.MultiIndex.from_arrays(
                [df.columns, ["value"] * len(df.columns)],
                names=["Dataset", "metrics"],
            )
            logger.debug("Columns after MultiIndex conversion: %s", df.columns)

        if isinstance(df.columns, pd.MultiIndex) and len(df.columns) > 0:
            existing_metric_labels = set(df.columns.get_level_values(1))
            new_metric_labels = {key for key in metrics.keys() if isinstance(key, str)}
            renamed_columns = []

            for dataset_label, metric_label in df.columns:
                renamed_label = metric_label
                if isinstance(metric_label, str) and metric_label.startswith("best_"):
                    base_metric = metric_label[len("best_") :]
                    theory_best_label = f"theory_best_{base_metric}"
                    if (
                        theory_best_label in new_metric_labels
                        and theory_best_label not in existing_metric_labels
                    ):
                        renamed_label = theory_best_label
                elif isinstance(metric_label, str) and metric_label.startswith("upper_bound_"):
                    base_metric = metric_label[len("upper_bound_") :]
                    theory_best_label = f"theory_best_{base_metric}"
                    if (
                        theory_best_label in new_metric_labels
                        and theory_best_label not in existing_metric_labels
                    ):
                        renamed_label = theory_best_label
                renamed_columns.append((dataset_label, renamed_label))

            df.columns = pd.MultiIndex.from_tuples(
                renamed_columns,
                names=df.columns.names,
            )
            logger.debug("Columns after legacy auxiliary metric migration: %s", df.columns)

        # --- 2) Collect datasets and metric info ---
        if isinstance(df.columns, pd.MultiIndex) and len(df.columns) > 0:
            existing_datasets = set(df.columns.get_level_values(0))
            existing_metric_labels = set(df.columns.get_level_values(1))
        else:
            existing_datasets = set()
            existing_metric_labels = set()

        logger.debug("Existing datasets: %s", existing_datasets)
        logger.debug("Existing metric labels: %s", existing_metric_labels)

        # Track base metrics and which bases already have auxiliary columns in the sheet.
        existing_metric_bases = set()
        bases_with_best_existing = set()
        bases_with_theory_best_existing = set()

        for label in existing_metric_labels:
            if not isinstance(label, str):
                existing_metric_bases.add(label)
                continue

            if label.startswith("best_"):
                base = label[len("best_") :]
                existing_metric_bases.add(base)
                bases_with_best_existing.add(base)
            elif label.startswith("theory_best_"):
                base = label[len("theory_best_") :]
                existing_metric_bases.add(base)
                bases_with_theory_best_existing.add(base)
            elif label.startswith("upper_bound_"):
                base = label[len("upper_bound_") :]
                existing_metric_bases.add(base)
                bases_with_theory_best_existing.add(base)
            else:
                existing_metric_bases.add(label)

        # From current metrics dict: infer base metrics and which bases have auxiliary metrics this run.
        new_metric_bases = set()
        bases_with_best_new = set()
        bases_with_theory_best_new = set()

        for key in metrics.keys():
            if not isinstance(key, str):
                new_metric_bases.add(key)
                continue

            if key.startswith("best_"):
                base = key[len("best_") :]
                new_metric_bases.add(base)
                bases_with_best_new.add(base)
            elif key.startswith("theory_best_"):
                base = key[len("theory_best_") :]
                new_metric_bases.add(base)
                bases_with_theory_best_new.add(base)
            elif key.startswith("upper_bound_"):
                base = key[len("upper_bound_") :]
                new_metric_bases.add(base)
                bases_with_theory_best_new.add(base)
            else:
                new_metric_bases.add(key)

        all_datasets = existing_datasets | {dataset_name}
        all_metric_bases = existing_metric_bases | new_metric_bases

        sorted_datasets = sorted(all_datasets)
        sorted_metric_bases = sorted(all_metric_bases)

        logger.info("All datasets after update: %s", sorted_datasets)
        logger.info("All base metrics after update: %s", sorted_metric_bases)
        logger.debug("Bases with existing best_* columns: %s", bases_with_best_existing)
        logger.debug("Bases with existing theory_best_* columns: %s", bases_with_theory_best_existing)
        logger.debug("Bases with new best_* metrics: %s", bases_with_best_new)
        logger.debug("Bases with new theory_best_* metrics: %s", bases_with_theory_best_new)

        # Build metric labels per base:
        #   always base,
        #   best_base only if either previously existed OR is present in current metrics,
        #   theory_best_base only if either previously existed OR is present in current metrics.
        metric_level_labels = []
        for base in sorted_metric_bases:
            metric_level_labels.append(base)
            if base in bases_with_best_existing or base in bases_with_best_new:
                metric_level_labels.append(f"best_{base}")
            if base in bases_with_theory_best_existing or base in bases_with_theory_best_new:
                metric_level_labels.append(f"theory_best_{base}")
        logger.debug("Metric-level labels (with conditional auxiliary columns): %s", metric_level_labels)

        # Build full MultiIndex: (Dataset x metric_level_labels)
        cols = pd.MultiIndex.from_product(
            [sorted_datasets, metric_level_labels],
            names=["Dataset", "metrics"],
        )
        logger.debug("New MultiIndex columns: %s", cols)

        df = df.reindex(columns=cols)
        logger.debug("DataFrame shape after column reindex: %s", df.shape)

        # --- 3) Ensure row index includes this model and is sorted ---
        existing_models = set(df.index)
        all_models = existing_models | {model_name}
        sorted_models = sorted(all_models)
        logger.info("All models after update: %s", sorted_models)

        df = df.reindex(index=sorted_models)
        logger.debug("DataFrame shape after index reindex: %s", df.shape)

        # --- 4) Insert / update the row for this model & dataset ---
        logger.info("Updating row for model '%s'.", model_name)
        if any(value is not None and not isinstance(value, Number) for value in metrics.values()):
            df = df.astype(object)

        for base_metric in new_metric_bases:
            base_key = base_metric
            best_key = f"best_{base_metric}"
            theory_best_key = f"theory_best_{base_metric}"

            if base_key in metrics:
                value = metrics[base_key]
                logger.debug(
                    "Setting base metric '%s' for model='%s', dataset='%s': value=%s",
                    base_metric,
                    model_name,
                    dataset_name,
                    value,
                )
                df.loc[model_name, (dataset_name, base_metric)] = value

            if best_key in metrics:
                best_value = metrics[best_key]
                logger.debug(
                    "Setting best metric '%s' for model='%s', dataset='%s': value=%s",
                    best_key,
                    model_name,
                    dataset_name,
                    best_value,
                )
                df.loc[model_name, (dataset_name, best_key)] = best_value

            if theory_best_key in metrics:
                theory_best_value = metrics[theory_best_key]
                logger.debug(
                    "Setting theory-best metric '%s' for model='%s', dataset='%s': value=%s",
                    theory_best_key,
                    model_name,
                    dataset_name,
                    theory_best_value,
                )
                df.loc[model_name, (dataset_name, theory_best_key)] = theory_best_value

        # Name index and column levels
        df.index.name = "Model"
        df.columns = df.columns.set_names(["Dataset", "metrics"])
        logger.debug("Index name: %s, column level names: %s", df.index.name, df.columns.names)

        # --- 5) Write data to Excel with MultiIndex headers ---
        logger.info("Writing updated DataFrame to Excel at '%s'.", excel_path)
        with pd.ExcelWriter(excel_path, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name=sheet_name)
        logger.info("Base DataFrame written to Excel successfully.")

        # --- 6) Post-process with openpyxl:
        #         - merge dataset headers (top row) across their metrics
        #         - bold per-dataset, per-base-metric extrema across models
        #         - adjust column widths based on metric names
        logger.info("Post-processing Excel with openpyxl.")
        wb = load_workbook(excel_path)
        ws = wb[sheet_name]

        # For MultiIndex columns with 2 levels, to_excel writes:
        #   row 1: column level 0 (Dataset) + something in A1
        #   row 2: column level 1 (metrics) + something in A2
        #   row 3: index name ('Model') in A3
        #   row 4+: data rows
        header_rows = df.columns.nlevels  # 2
        first_data_row = header_rows + 2  # 4 for 2 levels
        first_data_col = 2                # column 1 (A) is the index (Model)
        logger.debug(
            "Header rows: %d, first_data_row: %d, first_data_col: %d",
            header_rows,
            first_data_row,
            first_data_col,
        )

        # 6a) Merge dataset header cells in row 1
        logger.info("Merging dataset header cells in row 1.")
        current_dataset = None
        dataset_start_col = None
        dataset_end_cols: list[int] = []

        for pos, (ds, metric_name) in enumerate(df.columns):
            excel_col = first_data_col + pos

            if ds != current_dataset:
                # close previous dataset merge
                if current_dataset is not None and dataset_start_col is not None:
                    logger.debug(
                        "Merging dataset header '%s' from col %d to %d.",
                        current_dataset,
                        dataset_start_col,
                        excel_col - 1,
                    )
                    ws.merge_cells(
                        start_row=1,
                        start_column=dataset_start_col,
                        end_row=1,
                        end_column=excel_col - 1,
                    )
                    dataset_end_cols.append(excel_col - 1)
                current_dataset = ds
                dataset_start_col = excel_col

        if current_dataset is not None and dataset_start_col is not None:
            last_excel_col = first_data_col + len(df.columns) - 1
            logger.debug(
                "Final dataset header merge for '%s' from col %d to %d.",
                current_dataset,
                dataset_start_col,
                last_excel_col,
            )
            ws.merge_cells(
                start_row=1,
                start_column=dataset_start_col,
                end_row=1,
                end_column=last_excel_col,
            )
            dataset_end_cols.append(last_excel_col)

        # Label top-left header cells
        ws.cell(row=1, column=1).value = "Dataset"
        ws.cell(row=2, column=1).value = "metrics"
        logger.debug("Set header labels in A1 ('Dataset') and A2 ('metrics').")

        # 6b) Add vertical separators between datasets
        logger.info("Adding vertical separators between dataset column groups.")
        separator = Side(style="medium")
        max_row = ws.max_row
        for boundary_col in dataset_end_cols[:-1]:
            for row in range(1, max_row + 1):
                cell = ws.cell(row=row, column=boundary_col)
                cell.border = Border(
                    left=cell.border.left,
                    right=separator,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                )

        # 6c) Adjust column widths based on metric names
        logger.info("Adjusting column widths based on metric names.")
        MIN_WIDTH = 14  # looks good even for short names like 'f1'
        PADDING = 4     # extra space around the text

        for pos, (_, metric_name) in enumerate(df.columns):
            excel_col = first_data_col + pos
            col_letter = get_column_letter(excel_col)

            metric_str = str(metric_name)
            desired_width = max(len(metric_str) + PADDING, MIN_WIDTH)

            existing_width = ws.column_dimensions[col_letter].width
            if existing_width is None:
                final_width = desired_width
            else:
                final_width = max(existing_width, desired_width)

            ws.column_dimensions[col_letter].width = final_width
            logger.debug(
                "Set width of column '%s' to %s (metric='%s').",
                col_letter,
                final_width,
                metric_str,
            )

        # 6d) Bold extremum per dataset for base metrics and best_* auxiliary metrics.
        logger.info(
            "Applying bold font to per-dataset extrema across models for base metrics and best_* columns "
            "(max by default, min for metrics in %s).",
            minimize_set,
        )
        metric_level_vals = df.columns.get_level_values("metrics")

        for base_metric in sorted_metric_bases:
            metric_labels_to_bold = [base_metric]
            best_metric_label = f"best_{base_metric}"
            if best_metric_label in metric_level_vals:
                metric_labels_to_bold.append(best_metric_label)

            for metric_label in metric_labels_to_bold:
                logger.debug(
                    "Processing bolding for metric '%s' using base metric '%s'.",
                    metric_label,
                    base_metric,
                )

                if metric_label not in metric_level_vals:
                    logger.debug(
                        "Metric '%s' has no column (possibly only auxiliary columns existed); skipping.",
                        metric_label,
                    )
                    continue

                minimize = base_metric in minimize_set
                logger.debug(
                    "Metric '%s' minimize=%s (will bold %s).",
                    metric_label,
                    minimize,
                    "min" if minimize else "max",
                )

                # df_metric: index=models, columns=datasets
                df_metric = df.xs(metric_label, axis=1, level="metrics")

                for ds in sorted_datasets:
                    if ds not in df_metric.columns:
                        continue

                    col_series = df_metric[ds]
                    non_na = col_series.dropna()
                    if non_na.empty:
                        logger.debug(
                            "Dataset '%s' for metric '%s' has only NA; skipping.",
                            ds,
                            metric_label,
                        )
                        continue

                    if minimize:
                        best_val = non_na.min()
                    else:
                        best_val = non_na.max()

                    best_models = col_series[col_series == best_val].index
                    logger.debug(
                        "Dataset '%s', metric '%s': best_val=%s, best_models=%s",
                        ds,
                        metric_label,
                        best_val,
                        list(best_models),
                    )

                    # Excel column for (ds, metric_label)
                    col_pos = df.columns.get_loc((ds, metric_label))
                    excel_col = first_data_col + col_pos

                    for m_model in best_models:
                        row_pos = df.index.get_loc(m_model)
                        excel_row = first_data_row + row_pos
                        cell = ws.cell(row=excel_row, column=excel_col)
                        cell.font = Font(bold=True)

        wb.save(excel_path)
        logger.info("Excel post-processing complete and file saved: '%s'.", excel_path)
